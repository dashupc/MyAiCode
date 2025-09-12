import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import re
import time
from docx import Document
from pdfminer.high_level import extract_text
from bs4 import BeautifulSoup
import xlrd  # 处理xls
import openpyxl  # 处理xlsx
import comtypes.client  # 处理doc和ppt/pptx（需要Windows系统）

class MultiFormatExtractorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("多格式文件内容提取工具 | V0.1")
        self.root.geometry("800x600")
        self.root.resizable(True, True)
        
        # 设置中文字体支持
        self.style = ttk.Style()
        self.style.configure("TLabel", font=("SimHei", 10))
        self.style.configure("TButton", font=("SimHei", 10))
        self.style.configure("TEntry", font=("SimHei", 10))
        # 定义加大按钮的样式
        self.style.configure("Large.TButton", font=("SimHei", 12, "bold"), padding=10)
        
        # 创建主框架并居中
        self.main_frame = ttk.Frame(root, padding="30")
        self.main_frame.pack(expand=True, fill=tk.BOTH)
        
        # 输入文件夹路径
        self.input_dir = tk.StringVar()
        # 子文件夹名称
        self.subfolder_name = tk.StringVar(value="提取结果")
        
        # 支持的文件格式
        self.supported_formats = {
            'word': ['doc', 'docx'],
            'excel': ['xls', 'xlsx'],
            'powerpoint': ['ppt', 'pptx'],
            'pdf': ['pdf'],
            'web': ['html'],
            'text': ['txt']
        }
        
        # 创建UI元素
        self.create_widgets()
    
    def create_widgets(self):
        # 标题
        title_label = ttk.Label(self.main_frame, text="多格式文件内容提取工具", font=("SimHei", 16, "bold"))
        title_label.grid(row=0, column=0, columnspan=3, pady=20)
        
        # 输入文件夹选择
        ttk.Label(self.main_frame, text="输入文件夹:").grid(row=1, column=0, sticky=tk.W, pady=10)
        
        input_entry = ttk.Entry(self.main_frame, textvariable=self.input_dir, width=60)
        input_entry.grid(row=1, column=1, sticky=tk.W, pady=10, padx=5)
        
        browse_input_btn = ttk.Button(self.main_frame, text="浏览...", command=self.browse_input)
        browse_input_btn.grid(row=1, column=2, pady=10, padx=5)
        
        # 输出文件夹名称设置
        ttk.Label(self.main_frame, text="输出文件夹:").grid(row=2, column=0, sticky=tk.W, pady=10)
        
        subfolder_entry = ttk.Entry(self.main_frame, textvariable=self.subfolder_name, width=60)
        subfolder_entry.grid(row=2, column=1, sticky=tk.W, pady=10, padx=5)
        
        # 支持的格式说明 - 加大字号
        supported_formats = []
        for fmt_list in self.supported_formats.values():
            supported_formats.extend(fmt_list)
        format_label = ttk.Label(self.main_frame, text=f"支持格式: {', '.join(supported_formats)}", 
                                font=("SimHei", 11, "bold"))
        format_label.grid(row=3, column=0, columnspan=3, sticky=tk.W, pady=10)
        
        # 状态显示区域
        ttk.Label(self.main_frame, text="处理状态:").grid(row=4, column=0, sticky=tk.NW, pady=10)
        
        self.status_text = tk.Text(self.main_frame, height=12, width=60)
        self.status_text.grid(row=4, column=1, columnspan=2, pady=10, padx=5, sticky=tk.NSEW)
        scrollbar = ttk.Scrollbar(self.main_frame, command=self.status_text.yview)
        scrollbar.grid(row=4, column=3, sticky=tk.NS)
        self.status_text.config(yscrollcommand=scrollbar.set, state=tk.DISABLED)
        
        # 处理按钮 - 加大尺寸
        process_btn = ttk.Button(
            self.main_frame, 
            text="开始批量提取", 
            command=self.process_files, 
            style="Large.TButton"
        )
        process_btn.grid(row=5, column=0, columnspan=3, pady=20)
        
        # 配置网格权重，使界面可缩放
        self.main_frame.columnconfigure(1, weight=1)
        self.main_frame.rowconfigure(4, weight=1)
        
        # 底部信息
        footer_label = ttk.Label(self.main_frame, text="保留原格式 | 文件名格式: 旧文件名+提取信息 | V0.1", font=("SimHei", 8))
        footer_label.grid(row=6, column=0, columnspan=3, pady=10)
    
    def browse_input(self):
        """选择输入的文件夹"""
        dir_path = filedialog.askdirectory(title="选择包含文件的文件夹")
        if dir_path:
            self.input_dir.set(dir_path)
    
    def log(self, message):
        """在状态区域显示消息"""
        self.status_text.config(state=tk.NORMAL)
        self.status_text.insert(tk.END, message + "\n")
        self.status_text.see(tk.END)  # 滚动到最后
        self.status_text.config(state=tk.DISABLED)
        self.root.update_idletasks()  # 刷新界面
    
    def get_file_type(self, filename):
        """判断文件类型"""
        ext = filename.lower().split('.')[-1] if '.' in filename else ''
        for type_name, extensions in self.supported_formats.items():
            if ext in extensions:
                return type_name, ext
        return None, ext
    
    def extract_text_from_file(self, file_path, file_type, ext):
        """根据文件类型提取文本内容"""
        try:
            if file_type == 'word':
                return self.extract_word_text(file_path, ext)
            elif file_type == 'excel':
                return self.extract_excel_text(file_path, ext)
            elif file_type == 'powerpoint':
                return self.extract_ppt_text(file_path, ext)
            elif file_type == 'pdf':
                return self.extract_pdf_text(file_path)
            elif file_type == 'web':
                return self.extract_html_text(file_path)
            elif file_type == 'text':
                return self.extract_txt_text(file_path)
            else:
                return None
        except Exception as e:
            self.log(f"提取文本时出错: {str(e)}")
            return None
    
    def extract_word_text(self, file_path, ext):
        """提取Word文档文本"""
        if ext == 'docx':
            doc = Document(file_path)
            full_text = []
            for para in doc.paragraphs:
                full_text.append(para.text)
            return '\n'.join(full_text)
        elif ext == 'doc':  # 需要Windows系统支持
            word = comtypes.client.CreateObject('Word.Application')
            word.Visible = False
            doc = word.Documents.Open(os.path.abspath(file_path))
            content = doc.Content.Text
            doc.Close()
            word.Quit()
            return content
        return None
    
    def extract_excel_text(self, file_path, ext):
        """提取Excel文本"""
        text = []
        if ext == 'xlsx':
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text.append(f"工作表: {sheet_name}")
                for row in sheet.iter_rows(values_only=True):
                    row_text = [str(cell) if cell is not None else '' for cell in row]
                    text.append('\t'.join(row_text))
            workbook.close()
        elif ext == 'xls':
            workbook = xlrd.open_workbook(file_path)
            for sheet_idx in range(workbook.nsheets):
                sheet = workbook.sheet_by_index(sheet_idx)
                text.append(f"工作表: {sheet.name}")
                for row_idx in range(sheet.nrows):
                    row_text = [str(sheet.cell_value(row_idx, col_idx)) for col_idx in range(sheet.ncols)]
                    text.append('\t'.join(row_text))
        return '\n'.join(text)
    
    def extract_ppt_text(self, file_path, ext):
        """提取PowerPoint文本（需要Windows系统）"""
        try:
            powerpoint = comtypes.client.CreateObject('PowerPoint.Application')
            powerpoint.Visible = False
            presentation = powerpoint.Presentations.Open(os.path.abspath(file_path))
            
            text = []
            for slide_idx, slide in enumerate(presentation.Slides):
                text.append(f"幻灯片 {slide_idx + 1}")
                for shape in slide.Shapes:
                    if shape.HasTextFrame and shape.TextFrame.HasText:
                        text.append(shape.TextFrame.TextRange.Text)
            
            presentation.Close()
            powerpoint.Quit()
            return '\n'.join(text)
        except Exception as e:
            self.log(f"PPT提取错误: {str(e)}")
            return None
    
    def extract_pdf_text(self, file_path):
        """提取PDF文本"""
        try:
            return extract_text(file_path)
        except Exception as e:
            self.log(f"PDF提取错误: {str(e)}")
            return None
    
    def extract_html_text(self, file_path):
        """提取HTML文本"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                soup = BeautifulSoup(f.read(), 'html.parser')
                return soup.get_text()
        except Exception as e:
            self.log(f"HTML提取错误: {str(e)}")
            return None
    
    def extract_txt_text(self, file_path):
        """提取TXT文本"""
        try:
            encodings = ['utf-8', 'gbk', 'gb2312', 'latin-1']
            for encoding in encodings:
                try:
                    with open(file_path, 'r', encoding=encoding) as f:
                        return f.read()
                except UnicodeDecodeError:
                    continue
            # 如果所有编码都尝试失败
            with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                return f.read()
        except Exception as e:
            self.log(f"TXT提取错误: {str(e)}")
            return None
    
    def get_title_for_filename(self, file_path, file_type, ext):
        """从文件中提取标题作为文件名的一部分"""
        content = self.extract_text_from_file(file_path, file_type, ext)
        if not content:
            return "无内容"
        
        # 清理内容，提取前50个有效字符
        content = re.sub(r'\s+', ' ', content).strip()  # 替换多个空白为单个空格
        if len(content) > 50:
            return content[:50]
        return content if content else "无内容"
    
    def sanitize_filename(self, filename):
        """清理文件名中的非法字符"""
        invalid_chars = '/\\:*?"<>|'
        for char in invalid_chars:
            filename = filename.replace(char, '_')
        # 限制文件名长度
        return filename[:150]
    
    def process_files(self):
        """批量处理文件夹中的文件"""
        input_dir = self.input_dir.get()
        subfolder_name = self.subfolder_name.get()
        
        # 验证输入
        if not input_dir:
            messagebox.showerror("错误", "请选择输入文件夹")
            return
        
        if not os.path.exists(input_dir):
            messagebox.showerror("错误", "所选输入文件夹不存在")
            return
        
        if not subfolder_name.strip():
            messagebox.showerror("错误", "请输入子文件夹名称")
            return
        
        # 创建输出子文件夹
        output_dir = os.path.join(input_dir, subfolder_name)
        try:
            os.makedirs(output_dir, exist_ok=True)
        except Exception as e:
            messagebox.showerror("错误", f"无法创建输出子文件夹: {str(e)}")
            return
        
        # 获取文件夹中所有支持的文件
        all_files = []
        for f in os.listdir(input_dir):
            file_path = os.path.join(input_dir, f)
            if os.path.isfile(file_path):
                file_type, ext = self.get_file_type(f)
                if file_type:
                    all_files.append((f, file_type, ext))
        
        if not all_files:
            supported_formats = []
            for fmt_list in self.supported_formats.values():
                supported_formats.extend(fmt_list)
            messagebox.showinfo("提示", f"所选文件夹中没有找到支持的文件格式\n支持格式: {', '.join(supported_formats)}")
            return
        
        # 开始批量处理
        self.log(f"开始批量处理，共发现 {len(all_files)} 个支持的文件")
        self.log(f"提取结果将保存至: {output_dir}")
        success_count = 0
        error_count = 0
        
        for file, file_type, ext in all_files:
            file_path = os.path.join(input_dir, file)
            self.log(f"\n处理文件: {file}")
            
            # 获取原文件名（不含扩展名）
            original_name = os.path.splitext(file)[0]
            
            # 获取提取的信息作为文件名的一部分
            extracted_info = self.get_title_for_filename(file_path, file_type, ext)
            
            if not extracted_info:
                self.log(f"处理 {file} 失败：无法提取信息")
                error_count += 1
                continue
            
            # 组合旧文件名和提取的信息作为新文件名
            new_filename_base = f"{original_name}_{extracted_info}"
            
            # 清理文件名并添加扩展名
            filename = self.sanitize_filename(new_filename_base) + f".{ext}"
            
            # 处理可能的文件名重复
            base_filename, ext = os.path.splitext(filename)
            counter = 1
            output_path = os.path.join(output_dir, filename)
            
            # 如果文件已存在，添加编号
            while os.path.exists(output_path):
                filename = f"{base_filename}_{counter}{ext}"
                output_path = os.path.join(output_dir, filename)
                counter += 1
            
            # 复制文件（保留原格式）
            try:
                # 对于二进制文件直接复制
                if file_type in ['word', 'excel', 'powerpoint', 'pdf']:
                    with open(file_path, 'rb') as src, open(output_path, 'wb') as dst:
                        dst.write(src.read())
                # 对于文本文件，使用UTF-8编码保存
                else:
                    content = self.extract_text_from_file(file_path, file_type, ext)
                    if content:
                        with open(output_path, 'w', encoding='utf-8') as f:
                            f.write(content)
                    else:
                        raise Exception("无法提取文件内容")
                
                self.log(f"已保存至: {filename}")
                success_count += 1
                
                # 短暂延迟，避免界面卡顿
                self.root.update()
                time.sleep(0.1)
                
            except Exception as e:
                self.log(f"保存 {file} 时出错: {str(e)}")
                error_count += 1
        
        # 处理完成
        self.log(f"\n处理完成！成功: {success_count} 个, 失败: {error_count} 个")
        messagebox.showinfo("完成", f"批量处理完成！\n成功: {success_count} 个文件\n失败: {error_count} 个文件\n结果保存至: {output_dir}")

if __name__ == "__main__":
    # 创建自定义按钮样式
    root = tk.Tk()
    style = ttk.Style()
    style.configure("Accent.TButton", font=("SimHei", 10, "bold"))
    
    # 居中窗口
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry('{}x{}+{}+{}'.format(width, height, x, y))
    
    app = MultiFormatExtractorApp(root)
    root.mainloop()
    